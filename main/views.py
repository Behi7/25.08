from django.shortcuts import render, redirect
from . import models
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
import openpyxl


@login_required(login_url='login')
def listquiz(request):
    quizes = models.Quiz.objects.filter(author=request.user)
    context = {'quizes':quizes}
    return render(request, 'dashboard/listquiz.html', context)


def createquiz(request):
    if request.method == 'POST':
        models.Quiz.objects.create(
            name = request.POST['name'],
            author = request.user
    )
        return redirect('listquiz')
    return render(request, 'dashboard/createquiz.html')

def deletequiz(request, id):
    models.Quiz.objects.get(id=id).delete()
    return redirect('listquiz')

def listquestion(request, quiz_id):
    quiz = models.Quiz.objects.get(id=quiz_id)
    questions = models.Question.objects.filter(quiz=quiz)
    count_q = len(questions)
    context = {'questions':questions, 'quiz':quiz, 'count_q':count_q, 'count_a':len(models.Answer.objects.all())}
    return render(request, 'dashboard/listquestion.html', context)

def listoption(request, quiz_id, id):
    context = {}
    quiz_id = models.Quiz.objects.get(id=quiz_id)
    question = models.Question.objects.get(id=id)
    context['options'] = models.Option.objects.filter(question = question)
    context['quiz'] = quiz_id
    context['question'] = question
    return render(request, 'dashboard/listoption.html', context)

def createquestion(request, quiz_id):
    context = {}
    quiz = models.Quiz.objects.get(id=quiz_id)
    if request.method == "POST":
        question =  models.Question.objects.create(
            name = request.POST['name'],
            quiz = quiz
        )
        context['question'] = question
        return redirect('listquestion', quiz.id)
    context['quiz'] = quiz
    return render(request, 'dashboard/createquestion.html', context)

def deletequestion(request, quiz_id, id):
    quiz = models.Quiz.objects.get(id=quiz_id)
    context = {'quiz':quiz}
    models.Question.objects.get(id=id).delete()
    return redirect('listquestion', quiz_id=quiz_id)

def createoption(request, quiz_id, id):
    context = {}
    quiz = models.Quiz.objects.get(id=quiz_id)    
    question = models.Question.objects.get(id=id)
    context['correct'] = bool(models.Option.objects.filter(question = question))
    if request.method == 'POST':
        if not models.Option.objects.filter(question = question):
            models.Option.objects.create(
                name = request.POST['name'],
                question = question,
                correct = True
            )
            return redirect('listoption', quiz_id, question.id)
        else:
            models.Option.objects.create(
                name = request.POST['name'],
                question = question,
                correct = False
            )
            return redirect('listoption', quiz_id, question.id)

    context['quiz'] = quiz
    context['question'] = question
    return render(request, 'dashboard/createoption.html', context)

def log_out(request):
    logout(request)
    return redirect('listquiz')

def login_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            return redirect('listquiz')
        else:
            return redirect('login')
    return render(request, 'dashboard/login.html')

def register_user(request):
    context={}
    try:
        if request.method == 'POST':
            User.objects.create_user(
                username = request.POST['username'],
                password = request.POST['password'],
            )
            return redirect('login')
    except:
        context={'error':'bunday login bor'}
    return render(request, 'dashboard/register.html', context)

def deleteOption(request, quiz_id, question_id, id):
    models.Option.objects.get(id=id).delete()
    return redirect('listoption', quiz_id, question_id)

def updateoption(request, quiz_id, question_id, id):
    update = models.Option.objects.get(id=id)
    if request.method == 'POST':
        update.name = request.POST['name']
        update.save()
        return redirect('listoption', quiz_id, question_id)
    return render(request, 'dashboard/createoption.html', {'update':update, 'quiz_id':quiz_id, 'question_id':question_id})

def quizdetail(request, id):
    quiz = models.Quiz.objects.get(id=id)
    answers = models.Answer.objects.filter(quiz = quiz)
    for answer in answers:
        answerdetail = models.AnswerDetail.objects.filter(answer=answer)
        answer.options = len(answerdetail)
        filtered_correct = 0
        for obj in answerdetail:
            if obj.is_correct == True:
                filtered_correct+=1
        answer.un_correct = answer.options - filtered_correct
        answer.correct = filtered_correct
        if filtered_correct:
            answer.charget = (filtered_correct * 100) / answer.options
        else:
            answer.charget = 0
    return render(request, 'dashboard/quizdetail.html', {'answers':answers})

def renderxlsx(request, id):
    quiz = models.Quiz.objects.get(id=id)
    answers = models.Answer.objects.filter(quiz = quiz)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Ishlagan odam'
    sheet['B1'] = 'Test Nomi'
    sheet['C1'] = 'Savolar Soni'
    sheet['D1'] = 'Javoblar Soni'
    sheet['E1'] = "Tog'ri Javoblar"
    row = 2
    for answer in answers:
        answerdetail = models.AnswerDetail.objects.filter(answer=answer)
        answer.options = len(answerdetail)
        filtered_correct = 0
        for obj in answerdetail:
            if obj.is_correct == True:
                filtered_correct+=1
        answer.un_correct = answer.options - filtered_correct
        answer.correct = filtered_correct
        if filtered_correct:
            answer.charget = (filtered_correct * 100) / answer.options
        else:
            answer.charget = 0
        sheet.cell(row=row, column=1, value=answer.author.username)
        sheet.cell(row=row, column=2, value=quiz.name)
        sheet.cell(row=row, column=3, value=answer.options)
        sheet.cell(row=row, column=4, value=f"{answer.correct}/{answer.un_correct}")
        sheet.cell(row=row, column=5, value=f"{answer.charget}%")
        row += 1
    workbook.save(f'{quiz.name}.xlsx')
    return redirect('listquestion', quiz.id)

def answerxlsx(request, id):
    answer = models.Answer.objects.get(id=id)
    detail = models.AnswerDetail.objects.filter(answer = answer)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Savol'
    sheet['B1'] = 'Berilgan javob'
    sheet['C1'] = "to'g'ri javob"
    row = 2
    for item in detail:
        sheet.cell(row=row, column=1, value=item.question.name)
        sheet.cell(row=row, column=2, value=item.user_choice.name)
        sheet.cell(row=row, column=3, value=item.question.correct_option.name)
        row +=1
    workbook.save(f'{answer.author.username}.xlsx')
    return redirect('detailanswer', answer.id)